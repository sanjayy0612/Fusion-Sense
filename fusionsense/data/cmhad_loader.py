"""C-MHAD adapter: continuous third-person video + waist IMU -> FusionWindow.

Raw C-MHAD contains ten two-minute recordings per subject.  The spreadsheet
identifies each transition's recording number, class id, and start/end time.
Preparation crops one fixed two-second window centred on every annotation and
extracts the same real-time interval from both modalities.  Expensive
MediaPipe output is cached; model training reads only that compact cache.
"""
from __future__ import annotations

from collections import defaultdict
from dataclasses import replace
from pathlib import Path
import re

import numpy as np

from ..config import CFG, DATA_ROOT
from ..contract import ACTIVITIES, FusionWindow

PROCESSED_ROOT = DATA_ROOT.parent / "processed"
DEFAULT_CACHE = PROCESSED_ROOT / "cmhad_windows.npz"
EXPECTED_STREAM_SAMPLES = 6001


def find_transition_root(raw_root: str | Path | None = None) -> Path:
    """Locate the folder containing ``Subject1`` in an extracted release."""
    root = Path(raw_root or (DATA_ROOT / "cmhad")).expanduser().resolve()
    candidates = [root, root / "TransitionMovementsApplication"]
    candidates.extend(path.parent for path in root.glob("**/Subject1"))
    for candidate in candidates:
        if (candidate / "Subject1").is_dir():
            return candidate
    raise FileNotFoundError(
        f"Could not find C-MHAD TransitionMovementsApplication under {root}. "
        "Expected Subject1/...Subject12. See docs/DATASETS.md."
    )


def read_annotations(path: str | Path) -> list[dict]:
    """Read the official ``Video, Action, StartTime, EndTime`` spreadsheet."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ImportError("C-MHAD preparation needs openpyxl") from exc
    sheet = load_workbook(path, read_only=True, data_only=True).active
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value).strip().lower() for value in next(rows)]
    required = {
        "video": "video",
        "action": "action",
        "starttime(seconds)": "start",
        "endtime(seconds)": "end",
    }
    indices = {}
    for source, target in required.items():
        if source not in headers:
            raise ValueError(f"Missing {source!r} column in {path}; got {headers}")
        indices[target] = headers.index(source)

    parsed = []
    for row in rows:
        if not row or row[indices["video"]] is None:
            continue
        action = int(row[indices["action"]])
        if not 1 <= action <= len(ACTIVITIES):
            raise ValueError(f"Unexpected C-MHAD transition id {action} in {path}")
        parsed.append({
            "recording": int(row[indices["video"]]),
            "label": action - 1,
            "start": float(row[indices["start"]]),
            "end": float(row[indices["end"]]),
        })
    return parsed


def read_imu_stream(path: str | Path, expected_samples=EXPECTED_STREAM_SAMPLES) -> np.ndarray:
    """Read the official Shimmer CSV and restore its missing leading samples.

    The release documents a Bluetooth delay of roughly 30--40 samples.  Its
    reader restores alignment by left-padding each stream to 6001 samples.
    """
    array = np.genfromtxt(path, delimiter=",", skip_header=3, dtype=np.float32)
    if array.ndim != 2 or array.shape[1] < 7:
        raise ValueError(f"Expected timestamp + six IMU columns in {path}, got {array.shape}")
    values = array[:, 1:7]
    values = values[np.isfinite(values).all(axis=1)]
    missing = max(0, expected_samples - len(values))
    if missing:
        values = np.pad(values, ((missing, 0), (0, 0)), mode="constant")
    return values[:expected_samples].astype(np.float32)


def crop_stream_window(stream: np.ndarray, start_seconds: float, cfg=CFG) -> np.ndarray:
    """Crop/resample one aligned IMU interval to the contract shape."""
    first = int(round(start_seconds * cfg.imu_hz))
    indices = np.arange(first, first + cfg.t_imu)
    output = np.zeros((cfg.t_imu, cfg.imu_ch), dtype=np.float32)
    valid = (indices >= 0) & (indices < len(stream))
    output[valid] = stream[indices[valid], :cfg.imu_ch]
    return output


def _subject_number(path: Path) -> int:
    match = re.search(r"(\d+)$", path.name)
    if not match:
        raise ValueError(f"Cannot determine subject number from {path}")
    return int(match.group(1))


def prepare_cmhad(
    raw_root: str | Path | None = None,
    cache_path: str | Path = DEFAULT_CACHE,
    cfg=CFG,
    max_subjects: int | None = None,
) -> list[FusionWindow]:
    """Prepare all annotated transitions and save the compact training cache."""
    from .vision_extractor import video_to_aligned_pose_windows

    root = find_transition_root(raw_root)
    subject_dirs = sorted(root.glob("Subject*"), key=_subject_number)
    if max_subjects is not None:
        subject_dirs = subject_dirs[:max_subjects]
    windows = []

    for subject_dir in subject_dirs:
        subject = _subject_number(subject_dir)
        annotation_path = subject_dir / f"ActionOfInterestTraSubject{subject}.xlsx"
        annotations = read_annotations(annotation_path)
        by_recording = defaultdict(list)
        for annotation in annotations:
            by_recording[annotation["recording"]].append(annotation)

        for recording, events in sorted(by_recording.items()):
            events.sort(key=lambda event: event["start"])
            imu_path = subject_dir / "InertialData" / f"inertial_sub{subject}_tr{recording}.csv"
            video_path = subject_dir / "VideoData" / f"video_sub{subject}_tr{recording}.avi"
            if not imu_path.is_file() or not video_path.is_file():
                raise FileNotFoundError(f"Missing paired recording: {imu_path} / {video_path}")

            starts = [
                max(0.0, ((event["start"] + event["end"]) / 2.0) - cfg.window_seconds / 2.0)
                for event in events
            ]
            imu_stream = read_imu_stream(imu_path)
            pose_windows, valid_ratios, quality = video_to_aligned_pose_windows(
                str(video_path), starts, cfg=cfg
            )
            for event_index, (event, start) in enumerate(zip(events, starts)):
                windows.append(FusionWindow(
                    t_start=start,
                    imu=crop_stream_window(imu_stream, start, cfg),
                    radar=np.zeros((cfg.t_radar, cfg.radar_k), dtype=np.float32),
                    vision=pose_windows[event_index],
                    imu_valid=True,
                    radar_valid=False,
                    vision_valid=valid_ratios[event_index] >= 0.5,
                    radar_energy=0.0,
                    image_quality=quality[event_index],
                    imu_health=1.0,
                    label=event["label"],
                    subject_id=f"Subject{subject}",
                    recording_id=f"Subject{subject}/tr{recording}",
                ))
        print(f"prepared Subject{subject}: total windows={len(windows)}")

    if not windows:
        raise RuntimeError(f"No C-MHAD transitions were prepared from {root}")
    save_cmhad_windows(windows, cache_path, cfg)
    return windows


def save_cmhad_windows(windows, cache_path: str | Path = DEFAULT_CACHE, cfg=CFG):
    cache_path = Path(cache_path)
    cache_path.parent.mkdir(parents=True, exist_ok=True)
    np.savez_compressed(
        cache_path,
        version=np.array([1], dtype=np.int32),
        activities=np.array(ACTIVITIES),
        t_start=np.array([w.t_start for w in windows], dtype=np.float32),
        imu=np.stack([w.imu for w in windows]).astype(np.float32),
        vision=np.stack([w.vision for w in windows]).astype(np.float32),
        vision_valid=np.array([w.vision_valid for w in windows], dtype=bool),
        image_quality=np.array([w.image_quality for w in windows], dtype=np.float32),
        labels=np.array([w.label for w in windows], dtype=np.int64),
        subject_id=np.array([w.subject_id for w in windows]),
        recording_id=np.array([w.recording_id for w in windows]),
    )
    print(f"saved {len(windows)} synchronized windows -> {cache_path}")


def load_cmhad_windows(cache_path: str | Path = DEFAULT_CACHE, cfg=CFG) -> list[FusionWindow]:
    cache_path = Path(cache_path)
    if not cache_path.is_file():
        raise FileNotFoundError(
            f"Prepared C-MHAD cache not found at {cache_path}. Run "
            "`python scripts/prepare_cmhad.py --raw-root <path>` first."
        )
    data = np.load(cache_path, allow_pickle=False)
    cached_activities = list(data["activities"].astype(str))
    if cached_activities != ACTIVITIES:
        raise ValueError(
            f"C-MHAD cache labels {cached_activities} do not match code labels {ACTIVITIES}; "
            "re-run preparation."
        )
    windows = []
    for index in range(len(data["labels"])):
        windows.append(FusionWindow(
            t_start=float(data["t_start"][index]),
            imu=data["imu"][index].astype(np.float32),
            radar=np.zeros((cfg.t_radar, cfg.radar_k), dtype=np.float32),
            vision=data["vision"][index].astype(np.float32),
            imu_valid=True,
            radar_valid=False,
            vision_valid=bool(data["vision_valid"][index]),
            radar_energy=0.0,
            image_quality=float(data["image_quality"][index]),
            imu_health=1.0,
            label=int(data["labels"][index]),
            subject_id=str(data["subject_id"][index]),
            recording_id=str(data["recording_id"][index]),
        ))
    return windows


def fit_normalization(windows) -> dict[str, np.ndarray]:
    """Fit per-channel statistics on training subjects only."""
    stats = {}
    for name in ("imu", "vision"):
        selected = windows
        if name == "vision":
            selected = [window for window in windows if window.vision_valid]
            if not selected:
                raise ValueError("No MediaPipe-valid training windows available")
        values = np.concatenate([getattr(window, name) for window in selected], axis=0)
        mean = values.mean(axis=0).astype(np.float32)
        std = values.std(axis=0).astype(np.float32)
        stats[f"{name}_mean"] = mean
        stats[f"{name}_std"] = np.maximum(std, 1e-6)
    return stats


def normalize_windows(windows, stats) -> list[FusionWindow]:
    """Return copies normalized with training-only statistics."""
    return [
        replace(
            window,
            imu=((window.imu - stats["imu_mean"]) / stats["imu_std"]).astype(np.float32),
            vision=((window.vision - stats["vision_mean"]) / stats["vision_std"]).astype(np.float32),
        )
        for window in windows
    ]
